"""
excel_parser.py — reads TJ's original hand-maintained Osswaps workbook
(the "ON" sheet, in its native Client/Rep/Start/End/... layout) directly,
plus the "Transacted" sheet (TJ's own column layout), producing the
snapshot shape the app renders.

Dynamic rather than hardcoded to specific row numbers, since TJ edits
these sheets directly and row counts shift over time.
"""
from datetime import datetime, date
from io import BytesIO

import openpyxl


def _iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _swap_status(target, market, expiry_iso, today=None):
    today = today or date.today()
    expiry = None
    if expiry_iso:
        try:
            expiry = datetime.strptime(expiry_iso, "%Y-%m-%d").date()
        except ValueError:
            pass
    if expiry and expiry < today:
        return "Expired"
    if target is not None and market is not None and market < target:
        return "In the Money"
    return "Active"


def _standardize_product(label):
    if not label:
        return "Gas"
    l = str(label).strip().lower()
    if l.startswith("ab"):
        return "AB Power"
    if l.startswith("power"):
        return "Power"
    return "Gas"


def parse_deal_blocks(ws):
    swaps = []
    today = date.today()
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value != "Client":
            continue
        label_row = r - 1
        while label_row > 0 and not ws.cell(row=label_row, column=2).value:
            label_row -= 1
        product = _standardize_product(ws.cell(row=label_row, column=2).value)

        data_row = r + 1
        while ws.cell(row=data_row, column=2).value:
            client = ws.cell(row=data_row, column=2).value
            rep = ws.cell(row=data_row, column=3).value
            start = _iso(ws.cell(row=data_row, column=4).value)
            end = _iso(ws.cell(row=data_row, column=5).value)
            volume = _num(ws.cell(row=data_row, column=6).value)
            delivery = ws.cell(row=data_row, column=7).value
            target = _num(ws.cell(row=data_row, column=8).value)
            market = _num(ws.cell(row=data_row, column=9).value)
            expiry = _iso(ws.cell(row=data_row, column=10).value)
            sa = ws.cell(row=data_row, column=11).value

            status = _swap_status(target, market, expiry, today)
            days_left = None
            if expiry:
                try:
                    days_left = (datetime.strptime(expiry, "%Y-%m-%d").date() - today).days
                except ValueError:
                    pass

            swaps.append({
                "product": product,
                "client": str(client).strip() if client else "",
                "rep": rep,
                "start_date": start,
                "end_date": end,
                "volume": volume,
                "delivery_type": delivery,
                "target": target,
                "market": market,
                "expiry_date": expiry,
                "sa_number": sa,
                "transacted_date": None,
                "transacted_price": None,
                "savings_vs_target": None,
                "notes": None,
                "status": status,
                "days_to_expiry": days_left,
            })
            data_row += 1
    return swaps


def _infer_product_from_hub(hub):
    h = str(hub or "").strip().lower()
    if "7x24" in h or "5x16" in h:
        return "Power"
    if " lf" in h or h.endswith("lf") or "load following" in h:
        return "AB Power"
    return "Gas"


def parse_transacted_sheet(wb):
    """Reads TJ's actual Transacted sheet layout: header found dynamically
    (looks for 'Client' in column B), columns Client, Rep, Start, End,
    Volume, Del., Target, Transact, Transact Date, SA number, Savings vs
    Target — no Product column, inferred from delivery hub instead."""
    if "Transacted" not in wb.sheetnames:
        return []
    ws = wb["Transacted"]

    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        if ws.cell(row=r, column=2).value == "Client":
            header_row = r
            break
    if header_row is None:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        client = ws.cell(row=r, column=2).value
        if not client:
            continue
        rep = ws.cell(row=r, column=3).value
        start = _iso(ws.cell(row=r, column=4).value)
        end = _iso(ws.cell(row=r, column=5).value)
        volume = _num(ws.cell(row=r, column=6).value)
        delivery = ws.cell(row=r, column=7).value
        target = _num(ws.cell(row=r, column=8).value)
        transacted_price = _num(ws.cell(row=r, column=9).value)
        date_transacted = _iso(ws.cell(row=r, column=10).value)
        sa = ws.cell(row=r, column=11).value
        savings_vs_target = _num(ws.cell(row=r, column=12).value)

        rows.append({
            "product": _infer_product_from_hub(delivery),
            "client": str(client).strip(),
            "rep": rep,
            "start_date": start,
            "end_date": end,
            "volume": volume,
            "delivery_type": delivery,
            "target": target,
            "market": None,
            "expiry_date": None,
            "sa_number": sa,
            "transacted_date": date_transacted,
            "transacted_price": transacted_price,
            "savings_vs_target": savings_vs_target,
            "notes": None,
            "status": "Transacted",
            "days_to_expiry": None,
        })
    return rows


def parse_pricing_blocks(ws):
    header_rows = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=13).value
        start_cell = ws.cell(row=r, column=14).value
        end_cell = ws.cell(row=r, column=15).value
        p = ws.cell(row=r, column=16).value
        q = ws.cell(row=r, column=17).value
        rr = ws.cell(row=r, column=18).value
        looks_like_header = label and isinstance(p, str) and isinstance(q, str) and isinstance(rr, str)
        has_real_dates = isinstance(start_cell, (datetime, date)) or isinstance(end_cell, (datetime, date))
        if looks_like_header and not has_real_dates:
            header_rows.append((r, str(label).strip(), [p, q, rr]))

    sections = []
    for i, (hr, group_label, hubs) in enumerate(header_rows):
        boundary = header_rows[i + 1][0] - 1 if i + 1 < len(header_rows) else ws.max_row
        chunks = []
        current_chunk = []
        r = hr + 1
        while r <= boundary:
            period = ws.cell(row=r, column=13).value
            if period:
                start = _iso(ws.cell(row=r, column=14).value)
                end = _iso(ws.cell(row=r, column=15).value)
                prices = {}
                for j, hub in enumerate(hubs):
                    val = ws.cell(row=r, column=16 + j).value
                    if isinstance(val, (int, float)):
                        prices[hub] = f"${val:,.2f}"
                    elif val:
                        prices[hub] = str(val)
                    else:
                        prices[hub] = "—"
                current_chunk.append({"period_label": str(period).strip(), "start_period": start, "end_period": end, "prices": prices})
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                    current_chunk = []
            r += 1
        if current_chunk:
            chunks.append(current_chunk)

        if len(chunks) == 1:
            sections.append({"name": f"{group_label} Seasonal", "hubs": hubs, "rows": chunks[0]})
        else:
            names = [f"{group_label} Strip", f"{group_label} Seasonal"]
            for idx, chunk in enumerate(chunks):
                name = names[idx] if idx < len(names) else f"{group_label} {idx+1}"
                sections.append({"name": name, "hubs": hubs, "rows": chunk})

    return sections


def parse_workbook_bytes(xlsx_bytes):
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb["ON"]
    swaps = parse_deal_blocks(ws)
    swaps += parse_transacted_sheet(wb)
    pricing = parse_pricing_blocks(ws)
    return {
        "generated_at": datetime.now().isoformat(timespec="minutes"),
        "swaps": swaps,
        "pricing": pricing,
    }
